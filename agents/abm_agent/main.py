import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def run_abm_agent(
    df: pd.DataFrame,
    abm_df: pd.DataFrame,
    abm_filename: str,
    abm_type: str = "BNZSA QC",
    td_list: bool = False
) -> pd.DataFrame:
    """
    Compare company name and/or domain columns between df and abm_df.
    Fill the ABM file name into a single column:
    - First fully empty column from Custom 4 → Custom 18
    - If none empty, then first empty column among fallback columns
      (which exclude 'Additional Notes' if TD List mode is enabled)
    - If still none, default to Custom 18
    Fill 'No ABM' if no match.
    """

    # Lowercase copies for case-insensitive comparison
    df_lower = df.copy()
    df_lower.columns = df_lower.columns.str.strip().str.lower()
    abm_df_lower = abm_df.copy()
    abm_df_lower.columns = abm_df_lower.columns.str.strip().str.lower()

    # Prepare ABM data
# --- Combine multiple ABM lists safely ---
    abm_companies_series = abm_df_lower.get("company name", pd.Series(dtype=str)).dropna()
    abm_domains_series = abm_df_lower.get("domain", pd.Series(dtype=str)).dropna()

    # Remove duplicates and clean
    abm_companies = set(abm_companies_series.str.strip().str.lower())
    abm_domains = set(abm_domains_series.str.strip().str.lower())
    abm_company_map = {c.strip().lower(): c.strip() for c in abm_companies_series if str(c).strip()}

    print(f"📁 ABM combined total companies: {len(abm_companies)} | domains: {len(abm_domains)}")

    if not abm_companies and not abm_domains:
        print("⚠️ ABM list has no 'company name' or 'domain' column.")
        return df

    # Define column groups
    custom_cols = [f"Custom {i}" for i in range(4, 19)]
    fallback_cols = ["Comments", "Feedback For Publisher"]

    # Only include "Additional Notes" if TD list is NOT active
    if not td_list:
        fallback_cols.insert(0, "Additional Notes")

    # Ensure all possible columns exist
    for col in custom_cols + fallback_cols:
        if col not in df.columns:
            df[col] = ""

    # ---- Determine target column ----
    target_col = None
    for col in custom_cols:
        if df[col].isna().all() or (df[col].astype(str).str.strip() == "").all():
            target_col = col
            break

    if not target_col:
        for col in fallback_cols:
            if df[col].isna().all() or (df[col].astype(str).str.strip() == "").all():
                target_col = col
                break

    if not target_col:
        target_col = "Custom 18"

    print(f"✅ Using '{target_col}' as the single column for ABM output (TD list: {td_list})")

    # ---- Fill ABM filename ----
    def fill_abm_filename(row):
        company_lc = str(row.get("Company Name", "")).strip().lower()
        domain_lc = str(row.get("Domain", "")).strip().lower()
        matched = (company_lc in abm_company_map) or (domain_lc in abm_domains)
        if matched:
            return abm_filename if abm_type == "BNZSA QC" else "ABM"
        return "No ABM"

    df[target_col] = df.apply(fill_abm_filename, axis=1)

    # ---- Add ABM company name (if any matches exist) ----
    company_col_index = df.columns.get_loc("Company Name") if "Company Name" in df.columns else None
    if company_col_index is not None:
        if any(str(c).strip().lower() in abm_company_map for c in df["Company Name"].fillna("")):
            new_col = "ABM Company Name"
            if new_col not in df.columns:
                df.insert(company_col_index + 1, new_col, "")
            df[new_col] = df["Company Name"].apply(
                lambda c: abm_company_map.get(str(c).strip().lower(), "")
                if str(c).strip().lower() in abm_company_map else ""
            )

    print("✅ ABM mapping completed successfully.")
    return df


def apply_abm_coloring(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=2, values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and "ABM" in cell.value:
                cell.fill = highlight

    wb.save(filepath)
