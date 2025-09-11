import os, json, re, math
import pandas as pd
from typing import Tuple, Dict, Optional
from .fetcher import registered_domain
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

CACHE_FILE = "cache.json"

def load_cache(cache_path: str) -> Dict:
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_cache(cache_path: str, cache: Dict):
    tmp = cache_path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    os.replace(tmp, cache_path)

CANDIDATE_COMPANY_COLS = ["Company Name","company", "company_name", "name", "organization", "org"]
CANDIDATE_DOMAIN_COLS = ["Domain","domain", "website", "url", "site"]

def guess_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    cols = [c.lower().strip() for c in df.columns]
    company_col = None
    domain_col = None
    for c in CANDIDATE_COMPANY_COLS:
        for col in df.columns:
            if col.lower().strip() == c:
                company_col = col
                break
        if company_col:
            break
    for c in CANDIDATE_DOMAIN_COLS:
        for col in df.columns:
            if col.lower().strip() == c:
                domain_col = col
                break
        if domain_col:
            break
    return company_col, domain_col

def normalize_domain_cell(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in {"nan", "none", "null"}:
        return None
    # If a full URL, extract registered domain
    reg = registered_domain(s)
    return reg or s

def save_outputs(df: pd.DataFrame, out_csv: str, out_xlsx: Optional[str] = None):
    df.to_csv(out_csv, index=False, encoding="utf-8")
    if out_xlsx:
        df.to_excel(out_xlsx, index=False)  # Save plain Excel (no coloring)

# ---------------------------
# Enhanced coloring function
# ---------------------------
def apply_qc_domain_coloring(excel_file: str):
    """
    Apply QC Domain coloring to an existing Excel file based on 'match_status'.
    Color only 'match_status' and related columns, not the entire row.
    """
    RELATED_COLOR = "FFEB9C"       # Yellow
    MISMATCH_COLOR = "FFC7CE"      # Red
    INVALID_COLOR = "D9D9D9"       # Gray for invalid domains
    MANUAL_CHECK_COLOR = "ADD8E6"  # Light blue"

    RELATED_COLUMNS = [
        "match_status",
        "provided_company",
        "provided_domain",
        "discovered_domain",
        "discovered_company",
        "reason",
        "evidence_url",
    ]

    wb = load_workbook(excel_file)
    ws = wb.active

    # Map column names to their indices
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    if "match_status" not in col_map:
        print("⚠️ 'match_status' column not found, skipping coloring.")
        wb.save(excel_file)
        return

    for i in range(2, ws.max_row + 1):
        status_cell = ws.cell(row=i, column=col_map["match_status"])
        status_val = status_cell.value
        status = str(status_val).strip().lower() if status_val else ""

        # Decide color based on status
        fill_color = None
        if status == "match":
            fill_color = None  # No color for match
        elif status in {"related", "no_website"}:
            fill_color = RELATED_COLOR
        elif status == "mismatch":
            fill_color = MISMATCH_COLOR
        elif status == "invalid_domain":
            fill_color = INVALID_COLOR
        elif status == "need_manual_check":
            fill_color = MANUAL_CHECK_COLOR

        if fill_color:
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for col_name in RELATED_COLUMNS:
                col_idx = col_map.get(col_name)
                if col_idx:  # Only color if column exists in sheet
                    ws.cell(row=i, column=col_idx).fill = fill

    wb.save(excel_file)
