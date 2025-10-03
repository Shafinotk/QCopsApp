from openpyxl import load_workbook
from openpyxl.styles import PatternFill

LOW_VALUE_COLOR = "FFFF99"  # Light Yellow

def apply_irvalue_coloring(excel_file: str):
    """
    Apply coloring for IRValue results.
    Highlight rows where discovered_employees < 40 OR discovered_revenue < 10,000,000
    across these columns:
    discovered_employees_raw, discovered_revenue_raw, discovered_industry,
    flagged_rpe, discovered_employees, discovered_revenue.
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    # Map column names to indexes
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    target_cols = [
        "discovered_employees_raw",
        "discovered_revenue_raw",
        "discovered_industry",
        "flagged_rpe",
        "discovered_employees",
        "discovered_revenue",
    ]

    # Skip if required columns don't exist
    emp_col_idx = col_map.get("discovered_employees")
    rev_col_idx = col_map.get("discovered_revenue")
    if emp_col_idx is None and rev_col_idx is None:
        # Nothing to color, exit safely
        return

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        emp = ws.cell(row=i, column=emp_col_idx).value if emp_col_idx else None
        rev = ws.cell(row=i, column=rev_col_idx).value if rev_col_idx else None

        try:
            emp_val = float(emp) if emp is not None and emp != "" else None
            rev_val = float(rev) if rev is not None and rev != "" else None
        except ValueError:
            emp_val = None
            rev_val = None

        # Condition: highlight if emp < 40 OR rev < 10M
        if (emp_val is not None and emp_val < 40) or (rev_val is not None and rev_val < 10_000_000):
            fill = PatternFill(start_color=LOW_VALUE_COLOR, end_color=LOW_VALUE_COLOR, fill_type="solid")
            for col_name in target_cols:
                col_idx = col_map.get(col_name)
                if col_idx:  # Only fill if column exists
                    ws.cell(row=i, column=col_idx).fill = fill

    wb.save(excel_file)
