# agents/tollfree_agent/utils.py
import re
import pandas as pd
import openpyxl
import os
from openpyxl.styles import PatternFill

# Country aliases (normalize Excel country names → short codes used in dict)
COUNTRY_ALIASES = {
    "UNITED STATES": "US",
    "UNITED KINGDOM": "UK",
    "SOUTH KOREA": "KR",
    "NORTH KOREA": "KR",
    "RUSSIA": "RU",
    "CZECH REPUBLIC": "CZ",
    "CZECHIA": "CZ",
    "UAE": "AE",
    "UNITED ARAB EMIRATES": "AE",
    "SAUDI ARABIA": "SA",
    "SOUTH AFRICA": "ZA",
    "EGYPT": "EG",
    "NIGERIA": "NG",
    "AUSTRALIA": "AU",
    "INDIA": "IN",
    "JAPAN": "JP",
    "GERMANY": "DE",
    "FRANCE": "FR",
    "SPAIN": "ES",
    "SWEDEN": "SE",
    "CANADA": "CA",
    "MEXICO": "MX",
    # Add more mappings as needed...
}

COUNTRY_CODES = {}  # country code dictionary

# Default fallback prefixes
TOLL_FREE_PREFIXES = {
    "INTL": {"800"}
}

def load_extra_tollfrees():
    """
    Load toll-free prefixes and country codes from Excel file.
    Supports multiple country aliases in one cell (with or without quotes).
    """
    excel_path = os.path.join(os.path.dirname(__file__), "Countries_TollFree.xlsx")
    if not os.path.exists(excel_path):
        print(f"⚠️ Tollfree Excel not found: {excel_path}")
        return

    try:
        df = pd.read_excel(excel_path, dtype=str).fillna("")

        for _, row in df.iterrows():
            # --- Clean and split country aliases ---
            country_raw = str(row.get("Country", "")).replace("“", "").replace("”", "")
            country_aliases = []
            for c in country_raw.split(","):
                alias = c.strip().upper().replace('"', '').replace("'", "")
                if alias:
                    country_aliases.append(alias)

            # --- Clean and split toll-free prefixes ---
            prefixes_raw = str(row.get("Toll Free Numbers", "")).replace("“", "").replace("”", "")
            prefixes = []
            for p in prefixes_raw.split(","):
                prefix = re.sub(r"\D", "", p.replace('"', '').strip())
                if prefix:
                    prefixes.append(prefix)

            # --- Clean country code ---
            country_code_raw = (
                str(row.get("Country Codes", ""))
                .replace("“", "")
                .replace("”", "")
                .replace(" ", "")
                .replace('"', "")
            )
            country_code = country_code_raw if country_code_raw.isdigit() else None

            # --- Update dictionaries ---
            for alias in country_aliases:
                country = COUNTRY_ALIASES.get(alias, alias)
                if prefixes:
                    if country not in TOLL_FREE_PREFIXES:
                        TOLL_FREE_PREFIXES[country] = set()
                    TOLL_FREE_PREFIXES[country].update(prefixes)
                if country_code:
                    COUNTRY_CODES[country] = country_code

        # ✅ Debug print (optional)
        print("✅ Loaded toll-free prefixes:", TOLL_FREE_PREFIXES)

    except Exception as e:
        print(f"⚠️ Could not load toll-free prefixes: {e}")

# Run loader on import
load_extra_tollfrees()

def normalize_number(number: str) -> str:
    """Remove all non-digit characters from a phone number."""
    if not number or not isinstance(number, str):
        return ""
    return re.sub(r"\D", "", number)

def is_toll_free(number: str, country: str = None) -> bool:
    if not number:
        return False

    digits = normalize_number(number)
    country_key = COUNTRY_ALIASES.get(country.strip().upper(), country.strip().upper()) if country else None

    if not digits:
        return False

    # Country-specific prefixes
    if country_key and country_key in TOLL_FREE_PREFIXES:
        prefixes = TOLL_FREE_PREFIXES[country_key]
        for prefix in prefixes:
            if digits.startswith(prefix):
                return True

    # Fallback: check international tollfree prefixes
    for prefix in TOLL_FREE_PREFIXES.get("INTL", []):
        if digits.startswith(prefix):
            return True

    return False

def format_number(number: str, country: str = None, pattern: str = None) -> str:
    """
    Format phone number:
    - Use user-defined pattern if given
    - Keep existing special rules for India and USA
    - Otherwise: <country code> <rest of number>
    """
    digits = normalize_number(number)
    if not digits:
        return ""

    country_key = COUNTRY_ALIASES.get(country.strip().upper(), country.strip().upper()) if country else None
    country_code = COUNTRY_CODES.get(country_key)

    # --- User-defined pattern ---
    if pattern:
        digits_iter = iter(digits)
        formatted = ""
        for ch in pattern:
            if ch == "0":
                formatted += next(digits_iter, "")
            else:
                formatted += ch
        return formatted

    # --- Special rules for India and USA ---
    if country_key == "IN":
        if digits.startswith("91") and len(digits) == 12:
            return f"{digits[:2]} {digits[2:]}"  # 91 9876543210
        elif len(digits) == 10:
            return f"91 {digits}"  # Add country code if missing
    elif country_key in ("US", "CA"):
        if digits.startswith("1") and len(digits) == 11:
            return digits[1:]  # Strip leading 1
        elif len(digits) == 10:
            return digits  # Keep as 10 digits

    # --- Default for all other countries ---
    if country_code:
        if digits.startswith(country_code):
            digits = digits[len(country_code):]
        return f"{country_code} {digits}"

    # Fallback: return raw digits
    return digits

def apply_tollfree_coloring(excel_path: str, df: pd.DataFrame):
    """
    Color the 'Work Phone' column based on reason:
      - Toll-free numbers → Yellow
      - Other reasons (zeros, invalid length) → Light Red
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    tollfree_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Yellow
    other_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")    # Light Red

    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Work Phone":
            col_idx = idx
            break
    if col_idx is None:
        return

    for row_num, reason in enumerate(df.get("WorkPhone_Reason", []), 2):
        if reason == "tollfree":
            ws.cell(row=row_num, column=col_idx).fill = tollfree_fill
        elif reason == "other":
            ws.cell(row=row_num, column=col_idx).fill = other_fill

    wb.save(excel_path)
