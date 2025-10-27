import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def run_list_checker_agent(
    df: pd.DataFrame,
    competitor_df: pd.DataFrame | None = None,
    suppression_df: pd.DataFrame | None = None,
    td_df: pd.DataFrame | None = None,
    td_enabled: bool = False
) -> pd.DataFrame:
    """
    Check raw data against Competitor, Suppression, and TD lists.
    - Competitor/Suppression: mark rows with reason columns
    - TD: add 'TD Status Message' column if enabled and file uploaded

    Args:
        df: Raw dataframe
        competitor_df: Competitor list (may have company name, domain, email)
        suppression_df: Suppression list (may have company name, domain, email)
        td_df: TD list (must have 'email' and 'statusMessage')
        td_enabled: whether TD list mode is active
    """

    df = df.copy()
    
    # Ensure all list DataFrames are consistent
    if competitor_df is not None and isinstance(competitor_df, list):
        competitor_df = pd.concat(competitor_df, ignore_index=True)
    if td_df is not None and isinstance(td_df, list):
        td_df = pd.concat(td_df, ignore_index=True)


    # --- Normalize column names ---
    df.columns = df.columns.str.strip()

    # Helper: create lowercase sets for faster matching
    def to_set(series):
        return set(series.dropna().astype(str).str.strip().str.lower())

    # Initialize sets and maps
    comp_names, comp_domains, comp_emails = set(), set(), set()
    sup_names, sup_domains, sup_emails = set(), set(), set()
    td_map = {}

    # --- Competitor list ---
    if competitor_df is not None:
        cdf = competitor_df.copy()
        cdf.columns = cdf.columns.str.strip().str.lower()
        comp_names = to_set(cdf.get("company name", pd.Series(dtype=str)))
        comp_domains = to_set(cdf.get("domain", pd.Series(dtype=str)))
        comp_emails = to_set(cdf.get("email", pd.Series(dtype=str)))

    # --- Suppression list ---
    if suppression_df is not None:
        sdf = suppression_df.copy()
        sdf.columns = sdf.columns.str.strip().str.lower()
        sup_names = to_set(sdf.get("company name", pd.Series(dtype=str)))
        sup_domains = to_set(sdf.get("domain", pd.Series(dtype=str)))
        sup_emails = to_set(sdf.get("email", pd.Series(dtype=str)))

    # --- TD list ---
    if td_enabled and td_df is not None:
        tdf = td_df.copy()
        tdf.columns = tdf.columns.str.strip().str.lower()
        if "email" in tdf.columns and "statusmessage" in tdf.columns:
            td_map = dict(zip(
                tdf["email"].astype(str).str.lower(),
                tdf["statusmessage"]
            ))

    # --- Add reason columns if missing ---
    if competitor_df is not None and "Competitor_Reason" not in df.columns:
        df["Competitor_Reason"] = ""
    if suppression_df is not None and "Suppression_Reason" not in df.columns:
        df["Suppression_Reason"] = ""

    # --- Add TD column if enabled ---
    if td_enabled and td_df is not None:
        email_idx = df.columns.get_loc("Email") if "Email" in df.columns else len(df.columns) - 1
        if "TD Status Message" not in df.columns:
            df.insert(email_idx + 1, "TD Status Message", "")

    # --- Apply Competitor/Suppression/TD checks ---
    for idx, row in df.iterrows():
        cname = str(row.get("Company Name", "")).strip().lower()
        domain = str(row.get("Domain", "")).strip().lower()
        email = str(row.get("Email", "")).strip().lower()

        # Competitor Check
        if competitor_df is not None:
            if (cname in comp_names) or (domain in comp_domains) or (email in comp_emails):
                df.at[idx, "Competitor_Reason"] = "competitor list"

        # Suppression Check
        if suppression_df is not None:
            if (cname in sup_names) or (domain in sup_domains) or (email in sup_emails):
                df.at[idx, "Suppression_Reason"] = "suppression list"

        # TD Check
        if td_enabled and td_df is not None and "TD Status Message" in df.columns:
            if email in td_map:
                df.at[idx, "TD Status Message"] = td_map[email]

    # --- Drop internal boolean columns if they exist ---
    for col in ["Competitor List", "Suppression List"]:
        if col in df.columns:
            df.drop(columns=[col], inplace=True)

    print("✅ List Checker Agent completed successfully (with reason columns only).")
    return df


# ---------------- List Coloring for Excel ----------------
def apply_list_checker_coloring(filepath):
    wb = load_workbook(filepath)
    ws = wb.active
    comp_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    sup_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")
    td_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            val = str(cell.value).strip().lower()
            if val == "competitor list":
                cell.fill = comp_fill
            elif val == "suppression list":
                cell.fill = sup_fill
            elif val == "td list":
                cell.fill = td_fill

    wb.save(filepath)
