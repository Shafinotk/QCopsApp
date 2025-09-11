import re
import chardet
import pandas as pd
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =============================
# Constants & Configurations
# =============================

# Colors for Excel highlighting
EMAIL_MISMATCH_COLOR = "FFFF00"   # Yellow
RELATED_COLOR = "00B0F0"          # Blue (manual check needed)
MATCH_COLOR = "92D050"            # Green (valid match)
NOTSURE_COLOR = "FFA500"          # Orange (Not Sure cases)

# Generic domains to exclude
GENERIC_DOMAINS = {
    "gmail.com", "yahoo.com", "outlook.com",
    "icloud.com", "hotmail.com"
}

# =============================
# Helper Functions
# =============================

def clean_name(name: str) -> str:
    """Remove non-alphabetic characters and lowercase the string."""
    return re.sub(r"[^a-z]", "", name.lower().strip())

def has_middle_name(name: str) -> bool:
    """Check if a name string has more than one word (middle name or compound name)."""
    return len(name.split()) > 1

# =============================
# Main QC Function
# =============================

def qc_check(input_file: str, output_file: str, threshold: int = 70) -> str:
    """
    Perform QC checks on names, emails, and domains.
    Classifies rows into Match / Related / Mismatch / Invalid / Not Sure
    and applies Excel coloring for review.
    """

    # Detect encoding
    with open(input_file, "rb") as f:
        raw_data = f.read(100000)
        detected = chardet.detect(raw_data)
        encoding = detected["encoding"] or "utf-8"

    # Load CSV
    df = pd.read_csv(input_file, encoding=encoding)
    df.columns = [c.strip() for c in df.columns]

    # Add classification + reason columns
    df["email_status_Classification"] = ""
    df["Reason_for_name_mismatch_mail"] = ""
    df["Reason_for_domain_mismatch_mail"] = ""

    # =============================
    # Row-wise QC Checks
    # =============================
    for i, row in df.iterrows():
        first_name_raw = str(row["First Name"]).strip()
        last_name_raw = str(row["Last Name"]).strip()
        email = str(row["Email"]).strip().lower()
        domain = str(row["Domain"]).strip().lower()

        first_name = clean_name(first_name_raw)
        last_name = clean_name(last_name_raw)

        classification = ""
        name_reason = ""
        domain_reason = ""

        # -----------------------
        # Invalid Email Checks
        # -----------------------
        if "@" not in email or " " in email:
            classification = "Invalid format"
            name_reason = "Email missing '@' or contains spaces"
            df.at[i, "email_status_Classification"] = classification
            df.at[i, "Reason_for_name_mismatch_mail"] = name_reason
            continue

        email_user, email_domain = email.split("@", 1)

        if email_domain in GENERIC_DOMAINS:
            classification = "Invalid format"
            name_reason = f"Generic email domain used ({email_domain})"
            df.at[i, "email_status_Classification"] = classification
            df.at[i, "Reason_for_name_mismatch_mail"] = name_reason
            continue

        # -----------------------
        # Match Conditions
        # -----------------------
        if (
            first_name and last_name and
            (first_name in email_user and last_name in email_user)
        ):
            classification = "Match"

        elif last_name in email_user and email_user.startswith(first_name[0]):
            classification = "Match"

        elif first_name in email_user and email_user.endswith(last_name[0]):
            classification = "Match"

        elif has_middle_name(first_name_raw) or has_middle_name(last_name_raw):
            if first_name in email_user or last_name in email_user:
                classification = "Match"

        # -----------------------
        # Related Conditions
        # -----------------------
        elif email_user.endswith(last_name) and first_name not in email_user:
            classification = "Related"
            name_reason = "Email ends with last name; missing/unclear first name"

        elif email_user.startswith(first_name) and last_name not in email_user:
            classification = "Related"
            name_reason = "Email starts with first name; missing/unclear last name"

        elif (
            first_name and last_name and
            (
                (first_name[0] in email_user and last_name in email_user) or
                (last_name[0] in email_user and first_name in email_user)
            )
        ):
            classification = "Related"
            name_reason = "Initial + name combination; needs manual check"

        # -----------------------
        # Mismatch Cases
        # -----------------------
        elif re.fullmatch(r"[a-z]\.?[a-z]\.?[a-z]?", email_user):
            classification = "Mismatch"
            name_reason = "Email username looks like random initials"

        elif (
            (first_name and email_user.startswith(first_name[0])) or
            (last_name and email_user.startswith(last_name[0]))
        ):
            classification = "Mismatch"
            name_reason = "Only initials or weak match"

        # -----------------------
        # Not Sure Case
        # -----------------------
        if not classification:
            classification = "Not Sure"
            name_reason = "Pattern unclear, needs review"

        # -----------------------
        # Domain Check (Override if mismatch)
        # -----------------------
        if fuzz.ratio(domain, email_domain) < threshold:
            # If it was a "Match", downgrade to "Not Sure"
            if classification == "Match":
                classification = "Not Sure"
                domain_reason = "Check domain mismatch"
            else:
                domain_reason = "Domain mismatch with Email"

        # Save results
        df.at[i, "email_status_Classification"] = classification
        df.at[i, "Reason_for_name_mismatch_mail"] = name_reason
        df.at[i, "Reason_for_domain_mismatch_mail"] = domain_reason

    # =============================
    # Save to Excel with initial coloring
    # =============================
    df.to_excel(output_file, index=False)

    # Open for formatting
    wb = load_workbook(output_file)
    ws = wb.active
    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for i, row in df.iterrows():
        excel_row = i + 2
        class_val = row["email_status_Classification"]

        # Classification-based coloring
        if class_val == "Match":
            color = MATCH_COLOR
        elif class_val == "Not Sure":
            color = NOTSURE_COLOR
        elif "related" in class_val.lower():
            color = RELATED_COLOR
        elif "invalid" in class_val.lower() or "mismatch" in class_val.lower():
            color = EMAIL_MISMATCH_COLOR
        else:
            color = None

        if color:
            # Color Classification cell
            ws.cell(row=excel_row, column=col_map["email_status_Classification"]).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            # Color Email cell
            ws.cell(row=excel_row, column=col_map["Email"]).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            # Color Reason_for_name_mismatch_mail only if it has a value
            if row["Reason_for_name_mismatch_mail"]:
                ws.cell(row=excel_row, column=col_map["Reason_for_name_mismatch_mail"]).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            # Color Reason_for_domain_mismatch_mail only if it has a value
            if row["Reason_for_domain_mismatch_mail"]:
                ws.cell(row=excel_row, column=col_map["Reason_for_domain_mismatch_mail"]).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    wb.save(output_file)
    return output_file

# =============================
# New Function for Final Coloring
# =============================

def apply_mailname_coloring(excel_file: str):
    """
    Apply MailName coloring to an existing Excel file based on
    the 'email_status_Classification' column. Reasons inherit the same color.
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    col_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        class_val = ws.cell(row=i, column=col_map.get("email_status_Classification")).value
        color = None
        
        if class_val == "Not Sure":
            color = NOTSURE_COLOR
        elif class_val and "related" in str(class_val).lower():
            color = RELATED_COLOR
        elif class_val and ("invalid" in str(class_val).lower() or "mismatch" in str(class_val).lower()):
            color = EMAIL_MISMATCH_COLOR

        if color:
            ws.cell(row=i, column=col_map.get("email_status_Classification")).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            ws.cell(row=i, column=col_map.get("Email")).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )
            if ws.cell(row=i, column=col_map.get("Reason_for_name_mismatch_mail")).value:
                ws.cell(row=i, column=col_map.get("Reason_for_name_mismatch_mail")).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )
            if ws.cell(row=i, column=col_map.get("Reason_for_domain_mismatch_mail")).value:
                ws.cell(row=i, column=col_map["Reason_for_domain_mismatch_mail"]).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    wb.save(excel_file)
