import pandas as pd
import re
from collections import Counter
import openpyxl
from openpyxl.styles import PatternFill

# Direction tokens that should always be uppercase
DIRECTION_TOKENS = {
    "N", "S", "E", "W",
    "NE", "NS", "NW", "SN", "SE", "SW",
    "ES", "EW", "EN", "WE", "WS", "WN"
}

# Common street suffixes
STREET_TYPE_MAP = {
    "st": "St", "rd": "Rd", "ave": "Ave", "blvd": "Blvd",
    "ln": "Ln", "dr": "Dr", "hwy": "Hwy", "pkwy": "Pkwy",
    "cir": "Cir", "ct": "Ct", "pl": "Pl", "ter": "Ter", "way": "Way"
}

# Country and state short forms (kept uppercase)
COUNTRY_SHORTS = {
    "US", "USA", "GB", "UK", "IN", "IND", "CA", "CAN", "AU", "AUS", "NZ", "NZL",
    "AO", "AGO", "AT", "AUT", "AZ", "AZE", "BS", "BHS", "BH", "BHR", "BE", "BEL",
    "BO", "BOL", "BG", "BGR", "CL", "CHL", "CO", "COL", "CR", "CRI", "HR", "HRV",
    "CY", "CYP", "CZ", "CZE", "DK", "DNK", "DO", "DOM", "EC", "ECU", "EG", "EGY",
    "SV", "SLV", "EE", "EST", "FI", "FIN", "GR", "GRC", "HN", "HND", "HK", "HKG",
    "HU", "HUN", "IS", "ISL", "ID", "IDN", "IE", "IRL", "IL", "ISR", "JM", "JAM",
    "JO", "JOR", "KZ", "KAZ", "KE", "KEN", "KR", "KOR", "KP", "PRK", "KW", "KWT",
    "LV", "LVA", "LT", "LTU", "LU", "LUX", "MY", "MYS", "MU", "MUS", "MA", "MAR",
    "NG", "NGA", "NO", "NOR", "OM", "OMN", "PK", "PAK", "PA", "PAN", "PY", "PRY",
    "PE", "PER", "PH", "PHL", "PL", "POL", "PT", "PRT", "PR", "PRI", "QA", "QAT",
    "RO", "ROU", "RU", "RUS", "RS", "SRB", "SK", "SVK", "SI", "SVN", "LK", "LKA",
    "TW", "TWN", "TH", "THA", "TT", "TTO", "TN", "TUN", "TR", "TUR", "UG", "UGA",
    "UY", "URY", "VN", "VNM", "ZM", "ZMB", "CN", "CHN", "JP", "JPN", "DE", "DEU",
    "FR", "FRA", "IT", "ITA", "ES", "ESP", "BR", "BRA", "RU", "RUS", "MX", "MEX"
}


STATE_SHORTS = {
    "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA", "HI", "ID", "IL", "IN", "IA", "KS", "KY",
    "LA", "ME", "MD", "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ", "NM", "NY", "NC", "ND",
    "OH", "OK", "OR", "PA", "RI", "SC", "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY",
    "DL", "MP", "PB", "HR", "TN", "AP", "MH", "KA", "UP", "GJ", "RJ", "WB", "KL"  # Added few Indian codes
}

PUNCT_TO_REMOVE = r"[,\.\'\:\(\);]"
ORDINAL_RE = re.compile(r"(\d+)(st|nd|rd|th)", flags=re.IGNORECASE)
ZIPCODE_RE = re.compile(r"^\d+$")  # match only digits

# 🔑 Robust PO Box detection (supports P.O.Box, Pobox, Box 123, Post Office Box 123)
POBOX_RE = re.compile(
    r"\bP\.?\s*O\.?\s*Box\b|"       # P.O. Box
    r"\bPobox\b|"                   # Pobox
    r"\bPost\s*Office\s*Box\b|"     # Post Office Box
    r"\bBox\s*\d+\b",               # Box 123
    flags=re.IGNORECASE
)


def _canonical_col(df, target_name: str):
    tn = re.sub(r"\s+", "", target_name).lower()
    for c in df.columns:
        cc = re.sub(r"\s+", "", c).lower()
        if cc == tn:
            return c
    return None


def _clean_basic_text(value: str) -> str:
    if pd.isna(value):
        return ""
    value = str(value).strip()
    value = re.sub(PUNCT_TO_REMOVE, "", value)
    value = value.replace("-", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def clean_text(value: str) -> str:
    value = _clean_basic_text(value)
    return value.title() if value else ""


def _clean_country(value: str) -> str:
    """Keep short forms like US, IN, GB, CA (2/3-letter codes) in caps; else title-case."""
    value = _clean_basic_text(value)
    if not value:
        return ""
    upper_val = value.upper().strip()

    # Handle short codes
    if upper_val in COUNTRY_SHORTS:
        return upper_val

    # Handle long names with proper casing
    return value.title()



def _clean_state(value: str) -> str:
    """
    Clean the State column.
    - If it's 2 letters (like 'ny', 'ca', 'dl'), make it uppercase.
    - If it matches known short forms (STATE_SHORTS), keep uppercase.
    - Otherwise, title-case it.
    """
    value = _clean_basic_text(value)
    if not value:
        return ""

    upper_val = value.upper().strip()

    # If it's exactly 2 letters (short code), always uppercase
    if len(upper_val) == 2 or upper_val in STATE_SHORTS:
        return upper_val

    # Otherwise, title-case the name (e.g., 'California')
    return value.title()



def _propercase_street_token(token: str) -> str:
    if not token:
        return token
    upper = token.upper()

    if upper in DIRECTION_TOKENS:
        return upper

    if upper.lower() in STREET_TYPE_MAP:
        return STREET_TYPE_MAP[upper.lower()]

    m = ORDINAL_RE.fullmatch(token)
    if m:
        num, suf = m.groups()
        return f"{num}{suf.lower()}"

    return token.title()


def clean_street(value: str) -> str:
    """
    Clean the Street column.
    - Proper-case each word (street suffixes, directions).
    - Keep country short forms (like US, IN, UK, CA, etc.) in uppercase if they appear in the address.
    """
    value = _clean_basic_text(value)
    if not value:
        return value

    tokens = []
    for tok in value.split():
        upper_tok = tok.upper()
        # If token matches any known country code, keep it uppercase
        if upper_tok in COUNTRY_SHORTS:
            tokens.append(upper_tok)
        else:
            tokens.append(_propercase_street_token(tok))

    return " ".join(tokens)


def _properize_zip(zipcode: str, country: str) -> str:
    if pd.isna(zipcode):
        return ""
    zipcode = str(zipcode).strip()
    country = str(country).strip().upper()

    if country in ("US", "UNITED STATES", "USA", "UNITED STATES OF AMERICA", "United States", "United States Of America"):
        if ZIPCODE_RE.match(zipcode):
            if len(zipcode) == 4:
                return f"`0{zipcode}`"
            elif len(zipcode) == 5:
                return zipcode
            else:
                return f"INVALID:{zipcode}"
        else:
            return f"INVALID:{zipcode}"
    return zipcode


def _detect_pobox(street: str) -> bool:
    """Return True if PO Box is detected in the street"""
    if not street or pd.isna(street):
        return False
    return bool(POBOX_RE.search(street))


def apply_properization(df: pd.DataFrame, enforce_common_street: bool = True) -> pd.DataFrame:
    """Apply properization rules to DataFrame."""
    df = df.copy()

    mapping = {}
    for target in ["Company Name", "First Name", "Last Name", "Street", "City", "State", "Domain", "Country", "Zip Code"]:
        found = _canonical_col(df, target)
        if found:
            mapping[target] = found

    for col in ["Company Name", "First Name", "Last Name", "City"]:
        if col in mapping:
            c = mapping[col]
            df[c] = df[c].fillna("").astype(str).apply(clean_text)

    if "State" in mapping:
        c = mapping["State"]
        df[c] = df[c].fillna("").astype(str).apply(_clean_state)

    if "Country" in mapping:
        c = mapping["Country"]
        df[c] = df[c].fillna("").astype(str).apply(_clean_country)

    if "Street" in mapping:
        c = mapping["Street"]
        df["Street_POBox_Flag"] = df[c].fillna("").astype(str).apply(_detect_pobox)
        df[c] = df[c].fillna("").astype(str).apply(clean_street)

    if "Zip Code" in mapping and "Country" in mapping:
        zc = mapping["Zip Code"]
        cc = mapping["Country"]
        df[zc] = df.apply(lambda row: _properize_zip(row[zc], row[cc]), axis=1)

    if enforce_common_street and "Domain" in mapping:
        dom_col = mapping["Domain"]
        columns_to_enforce = ["Street", "City", "State", "Zip Code", "Country"]
        for col_name in columns_to_enforce:
            if col_name in mapping:
                col = mapping[col_name]
                chosen = {}
                for domain, group in df.groupby(dom_col):
                    non_empty = [s for s in group[col].astype(str) if s and s.strip()]
                    chosen_val = Counter(non_empty).most_common(1)[0][0] if non_empty else ""
                    chosen[domain] = chosen_val
                df[col] = df[dom_col].map(lambda d: chosen.get(d, ""))

    return df


def apply_pobox_coloring(excel_path: str, df: pd.DataFrame):
    """Color the 'Street' column if it contains PO Box (Yellow)."""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    col_idx = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Street":
            col_idx = idx
            break
    if col_idx is None:
        return

    for row_num, flag in enumerate(df.get("Street_POBox_Flag", []), 2):
        if flag:
            ws.cell(row=row_num, column=col_idx).fill = fill

    wb.save(excel_path)


if __name__ == "__main__":
    sample = pd.DataFrame({
        'Company Name': ['ACME, Inc.', 'acme inc'],
        'First Name': ['john', 'MARY'],
        'Last Name': ['doe', "o'neil"],
        'Street': [
            '123 N Main St.',
            'Pobox 79 Taman Perindustrian Lot 60334 Persiaran 3 Bukit Rahman Putra'
        ],
        'City': ['new york', 'NEW YORK'],
        'State': ['ny', 'California'],
        'Domain': ['acme.com', 'acme.com'],
        'Country': ['us', 'United States'],
        'Zip Code': ['2345', '123456']
    })
    result = apply_properization(sample, enforce_common_street=True)
    print(result)
